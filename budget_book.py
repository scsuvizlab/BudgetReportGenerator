"""
Enhanced budget spreadsheet parser with improved string analysis and contextual extraction.
Better handles names, positions, notes, and grant-specific content.
"""
import pandas as pd
import numpy as np
import re
import logging
from typing import Dict, List, Tuple, Any, Optional, Set, Union
from dataclasses import dataclass, field
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter
try:
    import xlrd
except ImportError:
    xlrd = None
try:
    from rapidfuzz import fuzz
except ImportError:
    fuzz = None

@dataclass
class BudgetCell:
    """Represents a cell in the budget with rich contextual information."""
    row: int
    col: int
    value: Any
    formatted_value: str
    data_type: str  # 'text', 'numeric', 'currency', 'date', 'formula'
    style_info: Dict[str, Any] = field(default_factory=dict)
    formula: Optional[str] = None
    hyperlink: Optional[str] = None
    comment: Optional[str] = None

@dataclass
class BudgetSection:
    """Represents a logical section within the budget."""
    name: str
    start_row: int
    end_row: int
    start_col: int
    end_col: int
    section_type: str  # 'personnel', 'equipment', 'travel', 'overhead', 'notes'
    cells: List[BudgetCell] = field(default_factory=list)
    metadata: Dict[str, Any] = field(default_factory=dict)

@dataclass
class PersonnelEntry:
    """Represents a personnel entry with name, title, and associated costs."""
    name: str
    title: str
    effort_percent: Optional[float] = None
    salary: Optional[float] = None
    fringe_rate: Optional[float] = None
    total_cost: Optional[float] = None
    notes: str = ""
    row_ref: int = -1

class BudgetBook:
    """Enhanced budget parser with improved string analysis and grant-specific parsing."""
    
    def __init__(self, enable_debug=False):
        self.enable_debug = enable_debug
        self.logger = logging.getLogger(__name__)
        
        # Regex patterns for various data types
        self.currency_patterns = [
            r'\$\s*[\d,]+\.?\d*',  # $1,234.56
            r'[\d,]+\.?\d*\s*\$',  # 1,234.56 $
            r'USD\s*[\d,]+\.?\d*',  # USD 1234
            r'[\d,]+\.?\d*\s*dollars?',  # 1234 dollars
        ]
        
        self.percentage_patterns = [
            r'[\d.]+\s*%',  # 25.5%
            r'[\d.]+\s*percent',  # 25.5 percent
        ]
        
        self.effort_patterns = [
            r'[\d.]+\s*%\s*effort',  # 25% effort
            r'[\d.]+\s*FTE',  # 0.25 FTE
            r'[\d.]+\s*fte',  # 0.25 fte
        ]
        
        # Keywords for identifying different budget sections
        self.section_keywords = {
            'personnel': {
                'headers': ['personnel', 'staff', 'salaries', 'wages', 'investigators', 'employees'],
                'indicators': ['name', 'title', 'effort', 'fte', 'salary', 'wage', 'pi', 'co-pi']
            },
            'equipment': {
                'headers': ['equipment', 'instruments', 'hardware', 'computers', 'software'],
                'indicators': ['item', 'description', 'quantity', 'unit cost', 'total cost']
            },
            'travel': {
                'headers': ['travel', 'conferences', 'meetings', 'transportation'],
                'indicators': ['destination', 'purpose', 'airfare', 'lodging', 'per diem', 'conference']
            },
            'supplies': {
                'headers': ['supplies', 'materials', 'consumables', 'reagents'],
                'indicators': ['description', 'quantity', 'unit', 'cost']
            },
            'overhead': {
                'headers': ['overhead', 'indirect', 'facilities', 'administrative', 'f&a'],
                'indicators': ['rate', 'base', 'total indirect']
            },
            'notes': {
                'headers': ['notes', 'justification', 'description', 'comments', 'explanation'],
                'indicators': ['note', 'describe', 'explain', 'justify']
            }
        }
        
        # Common title patterns for personnel
        self.title_patterns = [
            r'principal\s+investigator', r'pi\b', r'co-?pi\b',
            r'co-?investigator', r'research\s+(?:scientist|associate|professor)',
            r'post-?doc(?:toral)?(?:\s+(?:fellow|researcher))?',
            r'graduate\s+(?:student|research\s+assistant|assistant)',
            r'undergraduate\s+(?:student|research\s+assistant)',
            r'technician', r'specialist', r'coordinator', r'manager',
            r'professor', r'assistant\s+professor', r'associate\s+professor',
            r'lecturer', r'instructor', r'staff\s+(?:scientist|researcher)'
        ]
        
        # Name patterns (simple heuristics)
        self.name_indicators = [
            r'\b[A-Z][a-z]+\s+[A-Z][a-z]+\b',  # First Last
            r'\b[A-Z][a-z]+,\s*[A-Z][a-z]+\b',  # Last, First
            r'\b[A-Z][a-z]+\s+[A-Z]\.\s*[A-Z][a-z]+\b',  # First M. Last
            r'\b[A-Z]\.\s*[A-Z][a-z]+\b',  # F. Last
        ]
    
    def load_budget(self, file_path: Union[str, Path]) -> Dict[str, Any]:
        """Load and parse a budget file with enhanced string analysis."""
        file_path = Path(file_path)
        
        if not file_path.exists():
            raise FileNotFoundError(f"Budget file not found: {file_path}")
        
        self.logger.info(f"Loading budget file: {file_path}")
        
        # Determine file type and load accordingly
        if file_path.suffix.lower() in ['.xlsx', '.xlsm']:
            return self._load_excel_file(file_path)
        elif file_path.suffix.lower() == '.xls':
            return self._load_legacy_excel_file(file_path)
        elif file_path.suffix.lower() == '.csv':
            return self._load_csv_file(file_path)
        else:
            raise ValueError(f"Unsupported file format: {file_path.suffix}")
    
    def _load_excel_file(self, file_path: Path) -> Dict[str, Any]:
        """Load modern Excel file with rich formatting and context information."""
        workbook = openpyxl.load_workbook(file_path, data_only=False)
        budget_data = {
            'sheets': {},
            'metadata': {
                'filename': file_path.name,
                'sheet_names': workbook.sheetnames,
                'creation_date': getattr(workbook.properties, 'created', None),
                'last_modified': getattr(workbook.properties, 'modified', None)
            }
        }
        
        for sheet_name in workbook.sheetnames:
            self.logger.info(f"Processing sheet: {sheet_name}")
            worksheet = workbook[sheet_name]
            
            # Extract all cells with formatting and context
            sheet_cells = self._extract_excel_cells(worksheet)
            
            # Convert to DataFrame for analysis
            df = self._cells_to_dataframe(sheet_cells)
            
            # Analyze sheet structure and content
            sheet_analysis = self._analyze_sheet_structure(df, sheet_cells)
            
            # Extract personnel information
            personnel_entries = self._extract_personnel_entries(df, sheet_cells, sheet_analysis)
            
            # Extract budget sections
            budget_sections = self._extract_budget_sections(df, sheet_cells, sheet_analysis)
            
            budget_data['sheets'][sheet_name] = {
                'dataframe': df,
                'cells': sheet_cells,
                'analysis': sheet_analysis,
                'personnel': personnel_entries,
                'sections': budget_sections,
                'shape': (worksheet.max_row, worksheet.max_column)
            }
        
        workbook.close()
        return budget_data
    
    def _load_legacy_excel_file(self, file_path: Path) -> Dict[str, Any]:
        """Load legacy Excel file (.xls format)."""
        if xlrd is None:
            raise ImportError("Legacy Excel support requires 'xlrd' package. Install with: pip install xlrd")
            
        workbook = xlrd.open_workbook(file_path)
        budget_data = {
            'sheets': {},
            'metadata': {
                'filename': file_path.name,
                'sheet_names': workbook.sheet_names()
            }
        }
        
        for sheet_name in workbook.sheet_names():
            worksheet = workbook.sheet_by_name(sheet_name)
            
            # Convert to list of BudgetCell objects
            sheet_cells = []
            for row_idx in range(worksheet.nrows):
                for col_idx in range(worksheet.ncols):
                    cell = worksheet.cell(row_idx, col_idx)
                    if cell.value is not None:
                        budget_cell = BudgetCell(
                            row=row_idx,
                            col=col_idx,
                            value=cell.value,
                            formatted_value=str(cell.value),
                            data_type=self._determine_cell_type(cell.value)
                        )
                        sheet_cells.append(budget_cell)
            
            # Convert to DataFrame
            df = self._cells_to_dataframe(sheet_cells)
            
            # Analyze sheet
            sheet_analysis = self._analyze_sheet_structure(df, sheet_cells)
            personnel_entries = self._extract_personnel_entries(df, sheet_cells, sheet_analysis)
            budget_sections = self._extract_budget_sections(df, sheet_cells, sheet_analysis)
            
            budget_data['sheets'][sheet_name] = {
                'dataframe': df,
                'cells': sheet_cells,
                'analysis': sheet_analysis,
                'personnel': personnel_entries,
                'sections': budget_sections,
                'shape': (worksheet.nrows, worksheet.ncols)
            }
        
        return budget_data
    
    def _load_csv_file(self, file_path: Path) -> Dict[str, Any]:
        """Load CSV file with intelligent encoding detection."""
        # Try different encodings
        encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
        df = None
        
        for encoding in encodings:
            try:
                df = pd.read_csv(file_path, encoding=encoding, keep_default_na=False)
                break
            except UnicodeDecodeError:
                continue
        
        if df is None:
            raise ValueError(f"Could not read CSV file with any supported encoding: {file_path}")
        
        # Convert DataFrame to BudgetCell objects
        sheet_cells = []
        for row_idx in range(len(df)):
            for col_idx in range(len(df.columns)):
                value = df.iloc[row_idx, col_idx]
                if pd.notna(value) and str(value).strip():
                    budget_cell = BudgetCell(
                        row=row_idx + 1,  # Account for header row
                        col=col_idx,
                        value=value,
                        formatted_value=str(value),
                        data_type=self._determine_cell_type(value)
                    )
                    sheet_cells.append(budget_cell)
        
        # Analyze the single sheet
        sheet_analysis = self._analyze_sheet_structure(df, sheet_cells)
        personnel_entries = self._extract_personnel_entries(df, sheet_cells, sheet_analysis)
        budget_sections = self._extract_budget_sections(df, sheet_cells, sheet_analysis)
        
        return {
            'sheets': {
                'Sheet1': {
                    'dataframe': df,
                    'cells': sheet_cells,
                    'analysis': sheet_analysis,
                    'personnel': personnel_entries,
                    'sections': budget_sections,
                    'shape': df.shape
                }
            },
            'metadata': {
                'filename': file_path.name,
                'sheet_names': ['Sheet1']
            }
        }
    
    def _extract_excel_cells(self, worksheet) -> List[BudgetCell]:
        """Extract all cells from Excel worksheet with formatting information."""
        cells = []
        
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    # Extract style information
                    style_info = {
                        'font_name': cell.font.name if cell.font else None,
                        'font_size': cell.font.size if cell.font else None,
                        'font_bold': cell.font.bold if cell.font else False,
                        'font_italic': cell.font.italic if cell.font else False,
                        'fill_color': str(cell.fill.start_color.index) if cell.fill else None,
                        'number_format': cell.number_format,
                        'alignment': {
                            'horizontal': cell.alignment.horizontal if cell.alignment else None,
                            'vertical': cell.alignment.vertical if cell.alignment else None
                        }
                    }
                    
                    budget_cell = BudgetCell(
                        row=cell.row - 1,  # Convert to 0-based indexing
                        col=cell.column - 1,
                        value=cell.value,
                        formatted_value=str(cell.displayed_value) if hasattr(cell, 'displayed_value') else str(cell.value),
                        data_type=self._determine_cell_type(cell.value),
                        style_info=style_info,
                        formula=cell.formula if hasattr(cell, 'formula') else None,
                        hyperlink=cell.hyperlink.target if cell.hyperlink else None,
                        comment=cell.comment.text if cell.comment else None
                    )
                    cells.append(budget_cell)
        
        return cells
    
    def _cells_to_dataframe(self, cells: List[BudgetCell]) -> pd.DataFrame:
        """Convert BudgetCell objects to a pandas DataFrame."""
        if not cells:
            return pd.DataFrame()
        
        # Determine the dimensions
        max_row = max(cell.row for cell in cells) + 1
        max_col = max(cell.col for cell in cells) + 1
        
        # Create empty DataFrame
        data = [[None for _ in range(max_col)] for _ in range(max_row)]
        
        # Fill in the data
        for cell in cells:
            data[cell.row][cell.col] = cell.value
        
        # Convert to DataFrame
        df = pd.DataFrame(data)
        
        # Try to use first row as column names if they look like headers
        if len(df) > 0:
            first_row = df.iloc[0].dropna()
            if len(first_row) > 0 and all(isinstance(val, str) for val in first_row):
                # Check if first row looks like headers
                if self._looks_like_headers(first_row.tolist()):
                    df.columns = [f"Col_{i}" if i >= len(first_row) else str(first_row.iloc[i]) 
                                 for i in range(len(df.columns))]
                    df = df.iloc[1:].reset_index(drop=True)
        
        return df
    
    def _analyze_sheet_structure(self, df: pd.DataFrame, cells: List[BudgetCell]) -> Dict[str, Any]:
        """Analyze the structure of a sheet to identify sections and patterns."""
        analysis = {
            'total_rows': len(df),
            'total_cols': len(df.columns),
            'non_empty_cells': len(cells),
            'sections': {},
            'column_types': {},
            'potential_headers': [],
            'currency_cells': [],
            'text_cells': [],
            'formula_cells': [],
            'hyperlink_cells': [],
            'comment_cells': []
        }
        
        # Analyze column types
        for col_idx in range(len(df.columns)):
            col_data = df.iloc[:, col_idx].dropna()
            if not col_data.empty:
                analysis['column_types'][col_idx] = self._analyze_column_type(col_data)
        
        # Find potential sections based on keywords
        for section_type, keywords in self.section_keywords.items():
            section_ranges = self._find_section_ranges(df, keywords['headers'], keywords['indicators'])
            if section_ranges:
                analysis['sections'][section_type] = section_ranges
        
        # Categorize cells
        for cell in cells:
            if cell.data_type == 'currency':
                analysis['currency_cells'].append((cell.row, cell.col, cell.value))
            elif cell.data_type == 'text':
                analysis['text_cells'].append((cell.row, cell.col, cell.value))
            
            if cell.formula:
                analysis['formula_cells'].append((cell.row, cell.col, cell.formula))
            if cell.hyperlink:
                analysis['hyperlink_cells'].append((cell.row, cell.col, cell.hyperlink))
            if cell.comment:
                analysis['comment_cells'].append((cell.row, cell.col, cell.comment))
        
        # Find potential header rows
        analysis['potential_headers'] = self._find_header_rows(df)
        
        return analysis
    
    def _extract_personnel_entries(self, df: pd.DataFrame, cells: List[BudgetCell], 
                                  analysis: Dict[str, Any]) -> List[PersonnelEntry]:
        """Extract personnel entries with names, titles, and associated information."""
        personnel_entries = []
        
        # Look for personnel sections
        personnel_sections = analysis['sections'].get('personnel', [])
        
        for start_row, end_row, start_col, end_col in personnel_sections:
            # Extract personnel data from this section
            section_df = df.iloc[start_row:end_row+1, start_col:end_col+1]
            entries = self._parse_personnel_section(section_df, start_row, start_col)
            personnel_entries.extend(entries)
        
        # Also look for personnel patterns throughout the sheet
        general_entries = self._find_personnel_patterns_general(df, cells)
        personnel_entries.extend(general_entries)
        
        # Remove duplicates based on name similarity
        personnel_entries = self._deduplicate_personnel(personnel_entries)
        
        return personnel_entries
    
    def _extract_budget_sections(self, df: pd.DataFrame, cells: List[BudgetCell], 
                               analysis: Dict[str, Any]) -> List[BudgetSection]:
        """Extract structured budget sections from the spreadsheet."""
        sections = []
        
        for section_type, section_ranges in analysis['sections'].items():
            for start_row, end_row, start_col, end_col in section_ranges:
                section_cells = [cell for cell in cells 
                               if start_row <= cell.row <= end_row and start_col <= cell.col <= end_col]
                
                budget_section = BudgetSection(
                    name=f"{section_type.title()} Section",
                    start_row=start_row,
                    end_row=end_row,
                    start_col=start_col,
                    end_col=end_col,
                    section_type=section_type,
                    cells=section_cells,
                    metadata={
                        'cell_count': len(section_cells),
                        'has_currency': any(cell.data_type == 'currency' for cell in section_cells),
                        'has_formulas': any(cell.formula for cell in section_cells)
                    }
                )
                sections.append(budget_section)
        
        return sections
    
    def _determine_cell_type(self, value: Any) -> str:
        """Determine the type of data in a cell."""
        if pd.isna(value) or value is None:
            return 'empty'
        
        value_str = str(value).strip()
        
        # Check for currency
        if any(re.search(pattern, value_str) for pattern in self.currency_patterns):
            return 'currency'
        
        # Check for percentage
        if any(re.search(pattern, value_str) for pattern in self.percentage_patterns):
            return 'percentage'
        
        # Check for numeric
        if isinstance(value, (int, float)):
            return 'numeric'
        
        # Try to parse as number
        try:
            float(value_str.replace(',', '').replace('$', ''))
            return 'numeric'
        except ValueError:
            pass
        
        # Check for date patterns
        if self._looks_like_date(value_str):
            return 'date'
        
        # Default to text
        return 'text'
    
    def _looks_like_headers(self, values: List[str]) -> bool:
        """Check if a list of values looks like column headers."""
        if not values:
            return False
        
        # Headers are typically short, descriptive text
        avg_length = sum(len(str(val)) for val in values) / len(values)
        if avg_length > 30:  # Too long for headers
            return False
        
        # Should be mostly text
        text_count = sum(1 for val in values if isinstance(val, str))
        if text_count / len(values) < 0.7:
            return False
        
        # Check for common header words
        header_words = {'name', 'title', 'cost', 'amount', 'description', 'notes', 'total', 'rate', 'percent', 'fte'}
        found_header_words = sum(1 for val in values 
                               if isinstance(val, str) and 
                               any(hw in val.lower() for hw in header_words))
        
        return found_header_words > 0
    
    def _analyze_column_type(self, col_data: pd.Series) -> Dict[str, Any]:
        """Analyze the type and characteristics of a column."""
        analysis = {
            'primary_type': 'mixed',
            'numeric_ratio': 0.0,
            'text_ratio': 0.0,
            'currency_ratio': 0.0,
            'has_names': False,
            'has_titles': False,
            'avg_text_length': 0.0
        }
        
        total_count = len(col_data)
        if total_count == 0:
            return analysis
        
        numeric_count = 0
        text_count = 0
        currency_count = 0
        text_lengths = []
        
        for value in col_data:
            value_str = str(value).strip()
            
            if self._determine_cell_type(value) == 'currency':
                currency_count += 1
            elif self._determine_cell_type(value) == 'numeric':
                numeric_count += 1
            elif isinstance(value, str) and value.strip():
                text_count += 1
                text_lengths.append(len(value_str))
                
                # Check for names and titles
                if self._looks_like_person_name(value_str):
                    analysis['has_names'] = True
                if self._looks_like_job_title(value_str):
                    analysis['has_titles'] = True
        
        analysis['numeric_ratio'] = numeric_count / total_count
        analysis['text_ratio'] = text_count / total_count
        analysis['currency_ratio'] = currency_count / total_count
        analysis['avg_text_length'] = sum(text_lengths) / len(text_lengths) if text_lengths else 0
        
        # Determine primary type
        if analysis['currency_ratio'] > 0.5:
            analysis['primary_type'] = 'currency'
        elif analysis['numeric_ratio'] > 0.7:
            analysis['primary_type'] = 'numeric'
        elif analysis['text_ratio'] > 0.7:
            analysis['primary_type'] = 'text'
        
        return analysis
    
    def _find_section_ranges(self, df: pd.DataFrame, headers: List[str], 
                           indicators: List[str]) -> List[Tuple[int, int, int, int]]:
        """Find ranges that likely contain a specific type of section."""
        ranges = []
        
        for row_idx in range(len(df)):
            for col_idx in range(len(df.columns)):
                value = df.iloc[row_idx, col_idx]
                if isinstance(value, str):
                    value_lower = value.lower().strip()
                    
                    # Check if this cell contains a section header
                    if any(header in value_lower for header in headers):
                        # Found a potential section start, determine its extent
                        section_range = self._determine_section_extent(
                            df, row_idx, col_idx, indicators
                        )
                        if section_range:
                            ranges.append(section_range)
        
        # Merge overlapping ranges
        ranges = self._merge_overlapping_ranges(ranges)
        return ranges
    
    def _determine_section_extent(self, df: pd.DataFrame, start_row: int, start_col: int, 
                                indicators: List[str]) -> Optional[Tuple[int, int, int, int]]:
        """Determine the extent of a section starting from a header cell."""
        # Look for data in the next few rows that matches the indicators
        max_look_ahead = min(20, len(df) - start_row)
        
        # Find the extent of relevant data
        end_row = start_row
        end_col = start_col
        
        for row_offset in range(1, max_look_ahead):
            row_idx = start_row + row_offset
            has_relevant_data = False
            
            for col_idx in range(len(df.columns)):
                value = df.iloc[row_idx, col_idx]
                if isinstance(value, str):
                    value_lower = value.lower()
                    if any(indicator in value_lower for indicator in indicators):
                        has_relevant_data = True
                        end_row = max(end_row, row_idx)
                        end_col = max(end_col, col_idx)
                elif self._determine_cell_type(value) in ['currency', 'numeric', 'percentage']:
                    # Numeric data is often part of the section
                    has_relevant_data = True
                    end_row = max(end_row, row_idx)
                    end_col = max(end_col, col_idx)
            
            # If we hit several empty rows, stop
            if not has_relevant_data:
                empty_rows = 0
                for check_row in range(row_idx, min(row_idx + 3, len(df))):
                    if df.iloc[check_row].isna().all():
                        empty_rows += 1
                if empty_rows >= 2:
                    break
        
        # Ensure minimum section size
        if end_row - start_row < 1:
            return None
        
        return (start_row, end_row, start_col, min(end_col + 2, len(df.columns) - 1))
    
    def _merge_overlapping_ranges(self, ranges: List[Tuple[int, int, int, int]]) -> List[Tuple[int, int, int, int]]:
        """Merge overlapping or adjacent ranges."""
        if not ranges:
            return []
        
        # Sort by start row, then start column
        sorted_ranges = sorted(ranges, key=lambda x: (x[0], x[2]))
        merged = [sorted_ranges[0]]
        
        for current in sorted_ranges[1:]:
            last = merged[-1]
            
            # Check if ranges overlap or are adjacent
            if (current[0] <= last[1] + 1 and current[2] <= last[3] + 1):
                # Merge ranges
                merged[-1] = (
                    min(last[0], current[0]),
                    max(last[1], current[1]),
                    min(last[2], current[2]),
                    max(last[3], current[3])
                )
            else:
                merged.append(current)
        
        return merged
    
    def _find_header_rows(self, df: pd.DataFrame) -> List[int]:
        """Find rows that likely contain headers or section titles."""
        header_rows = []
        
        for row_idx in range(min(10, len(df))):  # Check first 10 rows
            row_data = df.iloc[row_idx].dropna()
            if not row_data.empty:
                # Calculate text ratio
                text_ratio = sum(1 for val in row_data if isinstance(val, str)) / len(row_data)
                
                # Check average text length (headers are usually short)
                text_values = [str(val) for val in row_data if isinstance(val, str)]
                avg_length = sum(len(val) for val in text_values) / len(text_values) if text_values else 0
                
                # Headers are mostly text with reasonable length
                if text_ratio > 0.6 and 3 <= avg_length <= 25:
                    header_rows.append(row_idx)
        
        return header_rows
    
    def _parse_personnel_section(self, section_df: pd.DataFrame, base_row: int, 
                               base_col: int) -> List[PersonnelEntry]:
        """Parse a personnel section to extract individual entries."""
        entries = []
        
        # Look for patterns indicating personnel entries
        for row_idx in range(len(section_df)):
            row_data = section_df.iloc[row_idx]
            
            # Look for names in this row
            name = None
            title = None
            salary = None
            effort = None
            notes = ""
            
            for col_idx, value in enumerate(row_data):
                if pd.isna(value):
                    continue
                
                value_str = str(value).strip()
                
                # Check if this looks like a name
                if self._looks_like_person_name(value_str):
                    name = value_str
                
                # Check if this looks like a title
                elif self._looks_like_job_title(value_str):
                    title = value_str
                
                # Check for salary/cost information
                elif self._determine_cell_type(value) == 'currency':
                    salary = self._extract_numeric_value(value_str)
                
                # Check for effort percentage
                elif any(re.search(pattern, value_str.lower()) for pattern in self.effort_patterns):
                    effort = self._extract_percentage_value(value_str)
                
                # Long text might be notes
                elif isinstance(value, str) and len(value_str) > 30:
                    notes = value_str
            
            # If we found a name, create an entry
            if name:
                entry = PersonnelEntry(
                    name=name,
                    title=title or "",
                    effort_percent=effort,
                    salary=salary,
                    notes=notes,
                    row_ref=base_row + row_idx
                )
                entries.append(entry)
        
        return entries
    
    def _find_personnel_patterns_general(self, df: pd.DataFrame, 
                                       cells: List[BudgetCell]) -> List[PersonnelEntry]:
        """Find personnel patterns throughout the entire sheet."""
        entries = []
        
        # Look for rows that contain names
        for row_idx in range(len(df)):
            row_data = df.iloc[row_idx]
            
            # Check if this row contains a person's name
            name_candidates = []
            for col_idx, value in enumerate(row_data):
                if isinstance(value, str) and self._looks_like_person_name(value):
                    name_candidates.append((col_idx, value))
            
            # For each name candidate, look for associated information
            for col_idx, name in name_candidates:
                entry = self._build_personnel_entry_from_row(df, row_idx, col_idx, name)
                if entry:
                    entries.append(entry)
        
        return entries
    
    def _build_personnel_entry_from_row(self, df: pd.DataFrame, row_idx: int, 
                                      name_col: int, name: str) -> Optional[PersonnelEntry]:
        """Build a personnel entry from a row containing a name."""
        title = None
        salary = None
        effort = None
        notes = ""
        
        # Look in nearby cells for related information
        row_data = df.iloc[row_idx]
        
        # Check cells in the same row
        for col_idx, value in enumerate(row_data):
            if col_idx == name_col or pd.isna(value):
                continue
            
            value_str = str(value).strip()
            
            if self._looks_like_job_title(value_str):
                title = value_str
            elif self._determine_cell_type(value) == 'currency':
                salary = self._extract_numeric_value(value_str)
            elif any(re.search(pattern, value_str.lower()) for pattern in self.effort_patterns):
                effort = self._extract_percentage_value(value_str)
            elif len(value_str) > 30:  # Likely notes
                notes = value_str
        
        # Also check the row above and below for titles or additional info
        for offset in [-1, 1]:
            check_row = row_idx + offset
            if 0 <= check_row < len(df):
                check_data = df.iloc[check_row]
                for col_idx in range(max(0, name_col - 2), min(len(check_data), name_col + 3)):
                    value = check_data.iloc[col_idx]
                    if isinstance(value, str) and self._looks_like_job_title(value):
                        if not title:  # Don't overwrite existing title
                            title = value
        
        return PersonnelEntry(
            name=name,
            title=title or "",
            effort_percent=effort,
            salary=salary,
            notes=notes,
            row_ref=row_idx
        )
    
    def _looks_like_person_name(self, text: str) -> bool:
        """Check if text looks like a person's name using heuristics."""
        if not isinstance(text, str) or len(text.strip()) < 3:
            return False
        
        text = text.strip()
        
        # Check against name patterns
        for pattern in self.name_indicators:
            if re.search(pattern, text):
                return True
        
        # Additional heuristics
        words = text.split()
        if len(words) < 2 or len(words) > 4:
            return False
        
        # Names typically start with capital letters
        if not all(word[0].isupper() for word in words if word):
            return False
        
        # Names shouldn't contain numbers or special characters (except periods, commas)
        if re.search(r'[0-9@#$%^&*()+=\[\]{}|\\:";\'<>?/]', text):
            return False
        
        # Check for common titles that aren't names
        title_words = {'dr', 'prof', 'professor', 'mr', 'mrs', 'ms', 'phd', 'md'}
        if any(word.lower() in title_words for word in words):
            return False
        
        return True
    
    def _looks_like_job_title(self, text: str) -> bool:
        """Check if text looks like a job title."""
        if not isinstance(text, str) or len(text.strip()) < 3:
            return False
        
        text_lower = text.lower().strip()
        
        # Check against title patterns
        for pattern in self.title_patterns:
            if re.search(pattern, text_lower):
                return True
        
        # Additional title indicators
        title_words = {
            'director', 'manager', 'coordinator', 'analyst', 'developer',
            'engineer', 'scientist', 'researcher', 'assistant', 'associate',
            'senior', 'junior', 'lead', 'chief', 'head', 'supervisor'
        }
        
        return any(word in text_lower for word in title_words)
    
    def _looks_like_date(self, text: str) -> bool:
        """Check if text looks like a date."""
        date_patterns = [
            r'\d{1,2}/\d{1,2}/\d{4}',  # MM/DD/YYYY
            r'\d{4}-\d{1,2}-\d{1,2}',  # YYYY-MM-DD
            r'\d{1,2}-\d{1,2}-\d{4}',  # MM-DD-YYYY
        ]
        
        return any(re.search(pattern, text) for pattern in date_patterns)
    
    def _extract_numeric_value(self, text: str) -> Optional[float]:
        """Extract numeric value from text (removing currency symbols, commas, etc.)."""
        if not isinstance(text, str):
            return None
        
        # Remove currency symbols and commas
        cleaned = re.sub(r'[$,\s]', '', text)
        
        try:
            return float(cleaned)
        except ValueError:
            return None
    
    def _extract_percentage_value(self, text: str) -> Optional[float]:
        """Extract percentage value from text."""
        if not isinstance(text, str):
            return None
        
        # Look for percentage patterns
        match = re.search(r'([\d.]+)\s*%', text)
        if match:
            try:
                return float(match.group(1))
            except ValueError:
                pass
        
        # Look for FTE patterns
        match = re.search(r'([\d.]+)\s*fte', text.lower())
        if match:
            try:
                return float(match.group(1)) * 100  # Convert FTE to percentage
            except ValueError:
                pass
        
        return None
    
    def _deduplicate_personnel(self, entries: List[PersonnelEntry]) -> List[PersonnelEntry]:
        """Remove duplicate personnel entries based on name similarity."""
        if not entries:
            return []
        
        if fuzz is None:
            # Simple deduplication without fuzzy matching
            unique_entries = []
            seen_names = set()
            for entry in entries:
                name_key = entry.name.lower().strip()
                if name_key not in seen_names:
                    unique_entries.append(entry)
                    seen_names.add(name_key)
            return unique_entries
        
        unique_entries = []
        
        for entry in entries:
            is_duplicate = False
            
            for existing in unique_entries:
                # Check name similarity
                similarity = fuzz.ratio(entry.name.lower(), existing.name.lower())
                if similarity > 85:  # High similarity threshold
                    is_duplicate = True
                    
                    # Keep the entry with more information
                    entry_info_count = sum([
                        bool(entry.title), bool(entry.salary), 
                        bool(entry.effort_percent), bool(entry.notes)
                    ])
                    existing_info_count = sum([
                        bool(existing.title), bool(existing.salary), 
                        bool(existing.effort_percent), bool(existing.notes)
                    ])
                    
                    if entry_info_count > existing_info_count:
                        # Replace existing with current entry
                        existing.title = entry.title or existing.title
                        existing.salary = entry.salary or existing.salary
                        existing.effort_percent = entry.effort_percent or existing.effort_percent
                        existing.notes = entry.notes or existing.notes
                    
                    break
            
            if not is_duplicate:
                unique_entries.append(entry)
        
        return unique_entries

    def get_all_text_values(self, budget_data: Dict[str, Any]) -> List[Tuple[str, str, Any]]:
        """Get all text values from the budget with their locations for field mapping."""
        text_values = []
        
        for sheet_name, sheet_data in budget_data['sheets'].items():
            cells = sheet_data['cells']
            
            for cell in cells:
                if cell.data_type == 'text' and isinstance(cell.value, str):
                    location = f"{sheet_name}!R{cell.row}C{cell.col}"
                    text_values.append((location, sheet_name, cell.value))
                
                # Also include comments and hyperlinks
                if cell.comment:
                    location = f"{sheet_name}!R{cell.row}C{cell.col}_comment"
                    text_values.append((location, sheet_name, cell.comment))
        
        return text_values

    def get_personnel_summary(self, budget_data: Dict[str, Any]) -> Dict[str, Any]:
        """Get a summary of all personnel found in the budget."""
        all_personnel = []
        
        for sheet_name, sheet_data in budget_data['sheets'].items():
            personnel = sheet_data.get('personnel', [])
            for person in personnel:
                person_dict = {
                    'sheet': sheet_name,
                    'name': person.name,
                    'title': person.title,
                    'effort_percent': person.effort_percent,
                    'salary': person.salary,
                    'notes': person.notes,
                    'row_ref': person.row_ref
                }
                all_personnel.append(person_dict)
        
        return {
            'total_personnel': len(all_personnel),
            'personnel_list': all_personnel,
            'unique_titles': list(set(p['title'] for p in all_personnel if p['title'])),
            'total_salary_budget': sum(p['salary'] for p in all_personnel if p['salary'])
        }